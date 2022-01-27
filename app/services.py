from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from itertools import islice
from datetime import datetime
import pandas as pd
import streamlit as st
import logging
import os

files = os.listdir('data')
workbooks = [item for item in files if '.xlsx' in item]

logging.basicConfig(filename='log.log', filemode='w', format='%(asctime)s - %(levelname)s %(message)s', datefmt='%H:%M:%S', encoding='utf-8', level=logging.DEBUG, force=True)

months = {'January': '01', 'February': '02', 'March': '03', 'April': '04', 'May': '05', 'June': '06', 'July': '07',
          'August': '08', 'September': '09', 'October': '10', 'November': '11', 'December': '12', 'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06', 'Jul': '07',
          'Aug': '08', 'Sep': '09', 'Oc': '10', 'Nov': '11', 'Dec': '12'}
years = ['2010','2011','2012','2013','2014','2015','2016','2017','2018','2019','2020','2021']

def check_file(file):
    """
    Checks if file exists and logs it.
    """
    while True:
        try:
            file = f'data/{file}'
            if os.path.exists(file):
                logging.info(f"{file} exists")
                break
        except FileNotFoundError:
            print("FileNotFound: not a valid file.")
            logging.warning("FileNotFound: not a valid file.")
        else:
            continue


def get_summary(ws, month_year_format):
    """
    Grabs relevant data from Summary MoM sheet.
    """
    row = None
    for item in ws['A']:
        if month_year_format in str(item.value):
            row = item.row
            st.write(f'(Row: {row})')
    values = [ro for ro in ws.iter_rows(min_row=row, max_row=row, values_only=True)]
    new_values = [item for item in values[0][1:] if item != None]

    # create dictionary from row data
    row_data = {}
    row_data['30s_abandonment'] = f'Abandon after 30s: {round(new_values[1]*100,2)}%'
    row_data['fcr'] = f'FCR : {new_values[2]*100}0%'
    row_data['dsat'] = f'DSAT : {new_values[3]*100}0%'
    row_data['csat'] = f'CSAT : {new_values[4]*100}0%'

    logging.info('get_summary succesful')

    return row_data


def nps_check(type, number):
    """
    Check size of group and return 'GOOD' or 'BAD'.
    """
    if type == 'promoters':
        if number >= 200:
            return 'GOOD'
        else:
            return 'BAD'
    if type == 'passives':
        if number >= 100:
            return 'GOOD'
        else:
            return 'BAD'
    if type == 'detractors':
        if number < 100:
            return 'GOOD'
        else:
            return 'BAD'
        
        
def get_voc(ws, month_year_format, month):
    """
    Grabs relevant data from VOC MoM sheet.
    """
    col = None
    
    try:
        for item in ws[1]:
            if month_year_format in str(item.value):
                col = item.column
                st.write(f'(Column: {col})')
    except ValueError:
        print('Month Year not found in VOC')
        logging.info('Month Year not found in VOC. Trying by month only...')
    else:
        for item in ws[1]:
            if month in str(item.value):
                col = item.column
                st.write(f'(Column: {col})')
            
    values = [co for co in ws.iter_cols(min_col=col, max_col=col, values_only=True)]
    new_values = [item for item in values[0][1:] if item != None and isinstance(item, int)]

    # create dictionary from column data
    col_data = {}
    col_data['base'] = f'Base Size: {new_values[0]}'
    col_data['promoters'] = [f'Promoters: {new_values[1]}', nps_check('promoters', new_values[1])]
    col_data['passives'] = [f'Passives: {new_values[2]}', nps_check('passives', new_values[2])]
    col_data['detractors'] = [f'Detractors: {new_values[3]}', nps_check('detractors', new_values[3])]

    logging.info('get_voc succesful')

    return col_data
    

def get_current():
    """
    Grabs the current month in integer / string formats and year.
    """
    # format month year for datetime comparison
    month = datetime.now().strftime('%m')
    month_word = datetime.now().strftime('%B')
    year = datetime.now().year

    logging.info(f'Current: {month_word}, {month}-{year}')

    return month, month_word, year
    


def log_summary(row_data):
    """
    Log Summary data.
    """
    print(row_data)
    for item in row_data:
        logging.info(row_data[item])


def log_voc(col_data):
    """
    Log VOC data.
    """
    for item in col_data:
        if 'base' in item:
            logging.info(col_data[item])
        else:
            logging.info(f'{col_data[item][0]} - {col_data[item][1]}')


def show_summary(row_data):
    """
    Display Summary data in streamlit app.
    """
    for item in row_data:
        st.write(row_data[item])

    logging.info(f'Displayed summary in app')


def show_voc(col_data):
    """
    Display VOC data in streamlit app.
    """
    for item in col_data:
        if 'base' in item:
            st.write(col_data[item])
        else:
            st.write(f'{col_data[item][0]} - {col_data[item][1]}')

    logging.info(f'Displayed voc in app')


def show_logs():
    with open('log.log') as log_file:
        for line in log_file:
            st.write(line)

    logging.info('Viewed logs')