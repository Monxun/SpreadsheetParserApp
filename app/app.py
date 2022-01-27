from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import logging
import os
import streamlit as st
import pandas as pd
from services import (
    files, 
    workbooks, 
    logging, 
    months, 
    years, 
    check_file,
    get_current, 
    get_summary, 
    get_voc,
    log_summary, 
    log_voc,
    show_summary, 
    show_voc,
    show_logs
)

#####################################################
#NAVIGATION

st.image('./app/spreadsheet-icon.png', width=100)
st.title('Spreadsheet Mini-Project')
pages = ['Select by File', 'Current Month', 'Log File']
navigation = st.selectbox('Select a page', pages)

st.write('_' * 30)

st.header(navigation)

data_flag = False

if navigation != 'Log File':

    if navigation == 'Select by File':

        #####################################################
        #SELECT FILE

        # workbook file selectbox 
        workbook_selector = st.selectbox('Select a file', workbooks)  # replace with streamlit selector 

        # verify file exists / error handling
        check_file(workbook_selector)

        # if file selected display text
        if workbook_selector:
            st.text(f'({workbook_selector} is selected.)')
            logging.info(f'({workbook_selector} is selected.)')

        # initialize workbook
        wb = load_workbook(f'data/{workbook_selector}')

        # grab month and year from file selection
        month_year = [item for item in workbook_selector.replace('.', '_').split("_") if item.capitalize() in months or item in years]

        # display month / year
        st.text(f'Month: {month_year[0].capitalize()}')
        st.text(f'Year: {month_year[1]}')

        # format month year for datetime comparison
        month_year_format = f'{month_year[1]}-{months[month_year[0].capitalize()]}'

        # data flag

        data_flag = True

        st.write('_' * 30)
        st.subheader(f'Data for {month_year[0].capitalize()} - {month_year[1]}')


    if navigation == 'Current Month':

        #####################################################
        #CURRENT MONTH

        data_flag = False

        month, month_word, year = get_current()
        st.write(f'{month_word}, {year}')
        current_files = []

        for book in workbooks:
            if month or month_word and year in book.replace('.', '_').split("_"):
                current_files.append(book)
                # data_flag = True
                

        if current_files:
            # initialize workbook
            wb = load_workbook(f'data/{current_files[0]}')

        st.subheader(f'Data for {month_word} - {year}')


    if navigation == 'Select by File' or data_flag == True:

        #####################################################
        #DATA

        # grab worksheet
        sheets = ['Summary Rolling MoM', 'VOC Rolling MoM']
        worksheet_selector = st.selectbox('Select a sheet', sheets)
        ws = wb[worksheet_selector]

        if worksheet_selector == 'Summary Rolling MoM':
            st.subheader('Summary')

            # get row data and return dictionary
            row_data = get_summary(ws, month_year_format)

            # log data
            log_summary(row_data)

            # show data
            show_summary(row_data)


        if worksheet_selector == 'VOC Rolling MoM':
            st.subheader('VOC')

            month = month_year[0].capitalize()

            # get column data and return dictionary
            col_data = get_voc(ws, month_year_format, month)

            # log data
            log_voc(col_data)

            # show data
            show_voc(col_data)

    else:
        st.subheader('NO MATCHING FILES')

else:

    #####################################################
    #LOG FILE

    show_logs()




